# encoding: utf-8
# encoding: iso-8859-1
# encoding: win-1252

from openpyxl import load_workbook
wb = load_workbook('Delivery.xlsx')
ws = wb['Delivery']

colProducto = 1
colEmpresa = 2
colTelefono = 5
colLocalizacion = 6
primerRow = 5


def buscarEmpresasVendenProducto(nombreProducto, primerRow=5, colProducto=1):
    contador = primerRow
    encontrado = False
    respuesta = ""
    while (ws.cell(row=contador, column=colProducto).value != None):
        if ws.cell(row=contador, column=colProducto).value == nombreProducto.upper():
            encontrado = True
            emp = ws.cell(row=contador, column=colEmpresa).value
            tel = ws.cell(row=contador, column=colTelefono).value
            loc = ws.cell(row=contador, column=colLocalizacion).value
            respuesta += emp + "\n\r"
            if tel != None:
                respuesta += u"Telefone/s: " + unicode(tel) + "\n\r"
            if loc != None:
                respuesta += u"Localização: " + unicode(loc) + "\n\r"
            respuesta += "\n\r"
        contador += 1
    if not encontrado:
        respuesta += u"O produto " + nombreProducto.upper()  + u" não aparece em nossos registros"
    return respuesta

#~ producto = raw_input(u"Digite o nome do produto: ")
#~ print buscarEmpresasVendenProducto(producto)
        
